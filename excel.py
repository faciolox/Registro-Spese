import openpyxl
'''
D2:D13 stipendio
E2:E13 altre entrate

H2:H13 TotaleSpese

D17:D28 SpeseVarie
E17:E28 Bonifico conto condiviso
F17:F28 Rata cofidis
G17:G28 Costi Fissi
H17:H28 Somma spese

L17:L28: Rata CC
M17:M28 Spese CC
N18:N28 Somma Spese CC

N2:N13 Saldo attuale all'8 del mese ( da febbraio a Gennaio dell'anno successivo )
'''
class Excel:
    """
    Classe Excel
    contenente workbook wb, utente, id, e precorso file
    """
    def __init__(self, utente, id):
        self.file, self.wb = carica_file(utente)
        self.utente = utente
        self.id = id

    def  getUtente(self):
        return self.utente
    def getId(self):
        return self.id
    def getFile(self):
        return self.file
    def getWb(self):
        return self.wb


diz_mesi = {
    'Tot':{1: 2, 2: 3, 3: 4, 4: 5, 5: 6, 6: 7, 7: 8, 8: 9, 9: 10, 10: 11, 11: 12, 12: 13},
     'Spese':{1: 17, 2: 18, 3: 19, 4: 20, 5: 21, 6: 22, 7: 23, 8: 24, 9: 25, 10: 26, 11: 27, 12: 28}
            }




def carica_file(utente):
    file = 'Registri/Riepilogo_Spese_' + utente  + '.xlsx'
    # carico il file excel
    wb = openpyxl.load_workbook(file, data_only=True)
    return file, wb

def carica_foglio(utente, anno):
    file = 'Registri/Riepilogo_Spese_' + utente  + '.xlsx'
    # carico il file excel
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb[anno]
    return sheet

def get_entrate_totali(anno,mese,wb):
    sheet = wb[anno]
    if mese == 0:
        mese = 12
        try:
            sheet = wb[int(anno)-1]
        except:
            return 0
    return sheet.cell(diz_mesi['Tot'][mese],4).value + sheet.cell(diz_mesi['Tot'][mese],5).value
def spesa_totale_mensile(anno, mese,wb):
    sheet = wb[anno]
    return sheet.cell(diz_mesi['Tot'][mese],8).value

def saldo_totale_mensile(anno, mese,wb):
    sheet = wb[anno]
    if mese == 0:
        mese = 12
        try:
            sheet = wb[int(anno)-1]
        except:
            return 0
    return sheet.cell(diz_mesi['Tot'][mese],14).value

def get_spese_varie( anno, mese, wb):
    sheet = wb[anno]
    return sheet.cell(diz_mesi['Spese'][mese],4).value

def get_bonifico_conto_condiviso( anno, mese, wb):
    sheet = wb[anno]
    return sheet.cell(diz_mesi['Spese'][mese], 5).value

def get_rata_cofidis( anno, mese, wb):
    sheet = wb[anno]
    return sheet.cell(diz_mesi['Spese'][mese], 6).value

def get_costi_fissi( anno, mese, wb):
    sheet = wb[anno]
    return sheet.cell(diz_mesi['Spese'][mese], 7).value

def get_somma_spese( anno, mese, wb):
    sheet = wb[anno]
    return sheet.cell(diz_mesi['Spese'][mese], 8).value

def get_rata_cc(anno,mese, wb):
    sheet = wb[anno]
    return sheet.cell(diz_mesi['Spese'][mese], 12).value

def get_spese_cc(anno,mese,wb):
    sheet = wb[anno]
    return sheet.cell(diz_mesi['Spese'][mese], 13).value

def get_somma_spese_cc(anno,mese,wb):
    sheet = wb[anno]
    return sheet.cell(diz_mesi['Spese'][mese], 14).value

def add_stipendio( anno, mese,stipendio, wb, file):
    sheet = wb[anno]
    sheet.cell(diz_mesi['Tot'][mese], 4, value=stipendio)
    wb.save(file)

def add_entrate( anno, mese,stipendio, wb, file):
    sheet = wb[anno]
    stipendio += sheet.cell(diz_mesi['Tot'][mese],5).value
    sheet.cell(diz_mesi['Tot'][mese], 5, value=stipendio)
    wb.save(file)

def add_spesa_varia( anno, mese, spesa, wb, file):
    sheet = wb[anno]
    spesa += sheet.cell(diz_mesi['Spese'][mese], 4).value
    sheet.cell(diz_mesi['Spese'][mese], 4, value=spesa)
    wb.save(file)

def add_bonifico_condiviso( anno, mese, spesa, wb, file):
    sheet = wb[anno]
    spesa += sheet.cell(diz_mesi['Spese'][mese], 5).value
    sheet.cell(diz_mesi['Spese'][mese], 5, value=spesa)
    wb.save(file)

def add_rata_cofidis( anno, mese, spesa, wb, file):
    sheet = wb[anno]
    spesa += sheet.cell(diz_mesi['Spese'][mese], 6).value
    sheet.cell(diz_mesi['Spese'][mese], 6, value=spesa)
    wb.save(file)

def add_spesa_cc( anno, mese, spesa, wb, file):
    sheet = wb[anno]
    spesa += sheet.cell(diz_mesi['Spese'][mese], 13).value
    sheet.cell(diz_mesi['Spese'][mese], 13, value=spesa)
    wb.save(file)

def add_rate_cc( anno, mese, spesa, mensilita, wb, file):
    sheet = wb[anno]
    rata_mensile = spesa / mensilita
    for i in range(mensilita+1):
        spesa = sheet.cell(diz_mesi['Spese'][mese + i], 12).value + rata_mensile
        sheet.cell(diz_mesi['Spese'][mese + i], 12, value=spesa)
    wb.save(file)

